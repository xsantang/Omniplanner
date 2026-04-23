import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\elxav\Downloads\Pago de cuentas mensualidades.xlsx", data_only=True)
print("Hojas:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    sep = "=" * 80
    print(f"\n{sep}")
    print(f"HOJA: {sn}  ({ws.max_row} filas x {ws.max_column} columnas)")
    print(sep)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False):
        vals = []
        for c in row:
            if c.value is not None:
                vals.append(f"{c.column_letter}{c.row}:{c.value}")
        if vals:
            print("  ".join(vals))
