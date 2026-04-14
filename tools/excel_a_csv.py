#!/usr/bin/env python3
"""
Convierte el Excel "Pago de cuentas mensualidades.xlsx" al formato CSV
que el rastreador de deudas de Omniplanner puede importar.

Formato de salida:
  cuenta,mes,saldo,pago,nuevos_cargos
  
  Donde 'nuevos_cargos' se calcula como:
    saldo_este_mes - (saldo_mes_anterior - pago_mes_anterior)
    Si es negativo o primer mes → 0
"""

import openpyxl
import csv
import sys
import os
from collections import defaultdict

EXCEL_PATH_DEFAULT = r"C:\Users\elxav\Downloads\Pago de cuentas mensualidades.xlsx"

# Cuentas a ignorar (son totales o entradas no-deuda)
IGNORAR = {
    "total", "total ", "total en el primer corte de pago",
    "total en el primer corte de pago ", "total en el segundo corte de pago",
    "total en el segundo corte de pago ", "canoochee emc",
}

# Mapeo de nombres del Excel → nombre normalizado
NORMALIZAR = {
    "att": "ATT",
    "bofa jennifer": "BOFA Jennifer",
    "discover": "Discover",
    "amazon": "Amazon",
    "amazon credit card": "Amazon",
    "usaa": "USAA",
    "navient": "Navient",
    "american express jennifer": "Amex Jennifer",
    "american express  xavico": "Amex Xavier",
    "american express xavico": "Amex Xavier",
    "american express": "Amex Xavier",
    "hyundai motor finance": "Hyundai",
    "bofa xavico": "BOFA Xavier",
    "bofa": "BOFA Xavier",
    "gci": "GCI",
    "coma i": "Coma I",
    "wyndham": "Wyndham",
    "whyham": "Wyndham",
    "dell": "Dell",
    "vacation club": "Wyndham",
    "vacation Club": "Wyndham",
    "best buy": "Best Buy",
    "geico": "GEICO",
    "firestone": "Firestone",
    "carro patty": "Carro Patty",
    "carro sra emma": "Carro Emma",
    "carro emma": "Carro Emma",
    "casa": "Casa",
    "pago de casa": "Casa",
    "seguro de la casa": "Seguro Casa",
    "victoria secret": "Victoria Secret",
    "reloj": "Reloj",
}

# Solo importar cuentas reconocidas (las que están en NORMALIZAR)
SOLO_CONOCIDAS = True

# Meses en español → número para ordenamiento
MES_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}

def normalizar_nombre(raw):
    """Normaliza el nombre de una cuenta. Retorna None si no es conocida."""
    key = raw.strip().lower()
    if SOLO_CONOCIDAS and key not in NORMALIZAR:
        return None
    return NORMALIZAR.get(key, raw.strip())

def es_cuenta_valida(nombre):
    """Determina si una fila es una cuenta real (no total, no header)."""
    key = nombre.strip().lower()
    if key in IGNORAR:
        return False
    if "quincena" in key or "corte" in key:
        return False
    if key.startswith("primera") or key.startswith("segunda"):
        return False
    # Nombres que son solo fechas
    if "de " in key and any(m in key for m in MES_NUM):
        return False
    return True

def extraer_mes_anio(sheet_name):
    """Extrae (mes_str, año) del nombre de hoja."""
    parts = sheet_name.strip().lower().split()
    if len(parts) == 1:
        # Solo mes (ej: "febrero") → asumir 2019
        mes = parts[0]
        return mes, 2019
    elif len(parts) == 2:
        mes, anio = parts[0], parts[1]
        try:
            return mes, int(anio)
        except ValueError:
            return mes, 2019
    return sheet_name.lower(), 2019

def sort_key(mes_anio):
    """Clave de ordenamiento para (mes, año)."""
    mes, anio = mes_anio
    return (anio, MES_NUM.get(mes, 0))

def main():
    # Aceptar ruta de entrada y salida como argumentos
    excel_path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH_DEFAULT
    output_override = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(excel_path):
        print(f"ERROR: No se encontró {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Estructura: {cuenta: [(mes_label, saldo, pago)]}
    datos = defaultdict(list)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        mes_str, anio = extraer_mes_anio(sheet_name)
        mes_label = f"{mes_str.capitalize()} {anio}"
        
        for row in ws.iter_rows(min_row=1, max_row=50, max_col=6, values_only=True):
            nombre_raw = row[0]
            if nombre_raw is None or not isinstance(nombre_raw, str):
                continue
            
            if not es_cuenta_valida(nombre_raw):
                continue
            
            saldo = row[1] if row[1] is not None else 0.0
            pago = row[2] if row[2] is not None else 0.0
            
            # Verificar que son números
            try:
                saldo = float(saldo)
                pago = float(pago)
            except (ValueError, TypeError):
                continue
            
            nombre = normalizar_nombre(nombre_raw)
            if nombre is None:
                continue
            datos[nombre].append((mes_str, anio, mes_label, saldo, pago))
    
    # Ordenar cada cuenta por fecha
    for cuenta in datos:
        datos[cuenta].sort(key=lambda x: (x[1], MES_NUM.get(x[0], 0)))
    
    # Consolidar: si una cuenta aparece 2+ veces en el mismo mes (quincenas),
    # tomar el saldo mayor como saldo_inicio y sumar los pagos
    datos_consolidados = {}
    for cuenta, registros in datos.items():
        consolidado = []
        grupo = defaultdict(list)
        for mes_str, anio, mes_label, saldo, pago in registros:
            grupo[(mes_str, anio, mes_label)].append((saldo, pago))
        
        for (mes_str, anio, mes_label), entradas in sorted(
            grupo.items(), key=lambda x: (x[0][1], MES_NUM.get(x[0][0], 0))
        ):
            # Saldo = el mayor encontrado (inicio del mes)
            saldo_max = max(e[0] for e in entradas)
            # Pago = suma de todos los pagos del mes
            pago_total = sum(e[1] for e in entradas)
            consolidado.append((mes_str, anio, mes_label, saldo_max, pago_total))
        
        datos_consolidados[cuenta] = consolidado
    
    # Calcular nuevos_cargos: si el saldo subió más de lo que indica 
    # (saldo_anterior - pago_anterior), la diferencia son nuevos cargos
    if output_override:
        output_path = output_override
    else:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..", "datos_deudas.csv"
        )
    
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["cuenta", "mes", "saldo", "pago", "nuevos_cargos"])
        
        total_filas = 0
        for cuenta in sorted(datos_consolidados.keys()):
            registros = datos_consolidados[cuenta]
            prev_saldo = None
            prev_pago = None
            
            for mes_str, anio, mes_label, saldo, pago in registros:
                nuevos_cargos = 0.0
                
                if prev_saldo is not None:
                    # Saldo esperado = anterior - pago anterior
                    esperado = prev_saldo - prev_pago
                    if saldo > esperado + 1.0:  # Margen de $1 por redondeo
                        nuevos_cargos = saldo - esperado
                
                writer.writerow([
                    cuenta,
                    mes_label,
                    f"{saldo:.2f}",
                    f"{pago:.2f}",
                    f"{nuevos_cargos:.2f}",
                ])
                total_filas += 1
                prev_saldo = saldo
                prev_pago = pago
    
    print(f"✓ Exportado: {os.path.abspath(output_path)}")
    print(f"  {len(datos_consolidados)} cuentas, {total_filas} registros mensuales")
    print(f"  Cuentas: {', '.join(sorted(datos_consolidados.keys()))}")

if __name__ == "__main__":
    main()
